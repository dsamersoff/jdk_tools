#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# vim: expandtab shiftwidth=2 softtabstop=2


_BANNER="""
  Simple BarCharter
  Version 1.000 2020-09-25
  Author: Dmitry Samersoff dms@samersoff.net
"""

import sys
import getopt

import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook

_logscale = False
_display = False

def read_xlsx(filename, sheetname):
    wb = load_workbook(filename=filename, data_only=True)
    if sheetname not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheetname}' not found in {filename}")
    ws = wb[sheetname]

    # Read labels
    labels = []
    for col in ws.iter_cols(min_col=2, min_row=1, max_row=1, values_only=True):
        if col[0] is None:
            break
        labels.append(col[0])

    # Read data
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        data.append((row[0],) + tuple(row[1:]))

    return labels, data

def usage(msg=None):
  global _HELP
  if msg is not None:
    print ("Error: %s" % msg)
  print(_HELP)
  sys.exit(7)

def main(args):
    filename = args[0]
    sheetname = args[1]

    labels, data = read_xlsx(filename, sheetname)

    print(20 * "-", "\n", repr(labels), "\n", 20 * "-")
    print(20 * "-", "\n", repr(data), "\n", 20 * "-")

    benchmarks = [row[0] for row in data]
    values = np.array([row[1:] for row in data], dtype=float)

    n_benchmarks = len(benchmarks)
    n_groups = values.shape[1]

    x = np.arange(n_benchmarks)  # positions
    width = 0.2  # bar width

    group_space = 0.3  # extra space between groups

    # make x positions with extra spacing
    x = np.arange(n_benchmarks) * (n_groups * width + group_space)

    fig, ax = plt.subplots(figsize=(14, 7))

    for i in range(n_groups):
        ax.bar(x + i*width, values[:, i], width, label=labels[i])

    ax.set_ylabel('Values')
    ax.set_title(sheetname)
    ax.set_xticks(x + width * (n_groups-1) / 2)
    ax.set_xticklabels(benchmarks, rotation=45, ha="right")
    ax.legend()

    if _logscale:
      plt.yscale("log")

    plt.tight_layout()
    plt.savefig("barchart_%s.png" % sheetname, dpi=300)

    if _display:
      plt.show()

if __name__ == "__main__":

  try:
    opts, args = getopt.getopt(sys.argv[1:],
                                "hVdl",
                               ["help", "version", "display", "logscale"])

    for o, a in opts:
      if o in ("-h", "--help", "-V", "--version"):
        print(_BANNER)
        if o in ("-h", "--help"):
          usage()
        sys.exit(7)
      elif o in ("-d", "--display"):
        _display = True 
      elif o in ("-l", "--logscale"):
        _logscale = True 
      else:
        assert False, "Unhandled option '%s'" % o

  # except getopt.GetoptError as ex:
  except Exception as ex:
    usage("Bad command line: %s (%s) " % (str(ex), repr(sys.argv[1:])))

  if len(args) != 2:
    usage()

  main(args)

