#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# vim: expandtab shiftwidth=2 softtabstop=2

# version 3.12 2020-05-05

HELP="""
  This is SPECJBB results gathering tool
  Usage: ./specjbb.py [-o outfile] [-s order] [-p filter] -t N -z dirname1 dirnameN
    -o outfile.xlsx | outfile.txt - dump all results and options to outfile
    -t N save only N topmoust results
    -z include zero-results run
    -s set sort order, might be prefix, name(default), max, crit
  """

import os
import sys
import getopt
import signal
import traceback
import codecs
import locale

import re
import fnmatch

try:
  import openpyxl
  from openpyxl import Workbook
  from openpyxl.styles import Alignment, PatternFill, Font

  _xls_possible = True
except ImportError as ex:
  _xls_possible = False

FILES = [ "*.raw", "*Backend*.gc.log", "Composite.gc.log" ]
README = "README.md"
OPTIONS = "options.txt"
SYSOPTS = "set_system.sh"

_full_dir_name = True
_out_name = None
_show_zeros = False
_sort = "name"
_top = None
_prefix_filter = "*"

_results = dict()
_cwd = os.getcwd()

""" Supporting function """
def signal_handler(signal, frame): #pylint: disable=unused-argument
  sys.stdout.write("\nInterrupted. Exiting ...\n")
  sys.exit(-1)

def report_exception(msg, ex):
  """ Report exception if requested """
  print(msg + "(%s)" % str(ex))
  print(traceback.format_exc())

"""  Result data class """
class ResultSet:
  MAX_PATTERN = re.compile(r"max-jOPS = ([0-9]+)")
  CRIT_PATTERN = re.compile(r"critical-jOPS = ([0-9]+)")
  
  # Pause Full (System.gc()) 235M->225M(228693M) 114.754ms
  # Pause Young (System.gc()) 4979M->235M(228693M) 3.153ms

  PAUSE_Y_PATTERN = re.compile(r"Pause Young.* ([0-9\.]+)ms")
  PAUSE_F_PATTERN = re.compile(r"Pause Full.* ([0-9\.]+)ms")
  
  RUN_NAME_PATTERN = re.compile(r"([A-Za-z0-9\-_]+\.)?[12][90]-[0-9]+-[0-9]+_[0-9]+")
 
  OPT_PATTERNS = { 
          "groups" : re.compile(r"^GROUP_COUNT: ([0-9]+)"),
          "pages" : re.compile(r"^PAGES: ([0-9]+)"),
          "SurvivorRatio" : re.compile(r"^JAVA_OPTS_BE: -XX:SurvivorRatio=([0-9]+)"),
          "TargetSurvivorRatio" : re.compile(r"^JAVA_OPTS_BE: -XX:TargetSurvivorRatio=([0-9]+)"),
          "MaxTenuringThreshold" : re.compile(r"^JAVA_OPTS_BE: -XX:MaxTenuringThreshold=([0-9]+)"),
  }

  def __init__ (self, prefix, run_name, filename):
    global _results, _cwd
    
    self.max_j = 0
    self.crit_j = 0
    self.max_gc_y = float(0)
    self.max_gc_f = float(0)
    self.filename = os.path.realpath(filename)
    self.run_name = run_name 
    self.sysopts_file = None
    self.options_file = None
    self.readme_file = None

    self.prefix = prefix
    if self.prefix == ".":
      self.prefix = _cwd.split(os.path.sep)[-1]

    (self.sysopts_file, self.options_file, self.readme_file) = \
      ResultSet.get_file_set(self.filename) 

    """ Self registration """
    _results[self.run_name] = self 
    
  @staticmethod
  def get_run_name(filename):
    """ Extract run name from the path to the file with results, assume yyyy-mm-dd_HHMMSS"""
    pi = filename.split(os.path.sep)
    pi_iter = reversed(pi)
    (prefix, run_name) = ("Amnesiac", "Amnesiac")
    for pi_item in pi_iter:
      if ResultSet.RUN_NAME_PATTERN.match(pi_item):
        run_name = pi_item
        prefix = next(pi_iter, "Amnesiac")
        break
    return (prefix, run_name)

  @staticmethod
  def get_file_set(filename):
    """ Go from filename directory upward until all required description files found """
    global SYSOPTS, OPTIONS, README, _cwd

    rdir = filename 
    rname = None
    (sysopts, options, readme) = (None, None, None)
    
    while (rdir != "" and rdir != _cwd):
      rdir = os.path.dirname(rdir)
      rname = os.path.join(rdir, SYSOPTS)
      if os.path.exists(rname) and sysopts is None:
        sysopts = rname
      rname = os.path.join(rdir, OPTIONS)
      if os.path.exists(rname) and options is None:
        options = rname
      rname = os.path.join(rdir, README)
      if os.path.exists(rname) and readme is None:
        readme = rname
    return (sysopts, options, readme)

  def read_description(self):
    """ Read README.md content, assume one line"""
    desc = ""
    if self.readme_file != None:
      with open(self.readme_file, "r") as fd:
        ln = fd.readline()
        desc = ln[:-1]
    return desc

  def read_options(self):
    """ Read and parse options.txt file, extract only options we are interesing in"""
    opts = dict()
    if self.options_file != None:
      with open(self.options_file, "r") as fd:
        for ln in fd:
          for (p_name, p_val) in ResultSet.OPT_PATTERNS.items():   
            m = p_val.search(ln)
            if m != None and m.group(1) != None:
              opts[p_name]  = m.group(1)
    return opts
  
  def set_max_crit(self, ln): 
    """ Search for max and crit scores """
    m = ResultSet.MAX_PATTERN.search(ln)
    if m != None and m.group(1) != None:
      self.max_j = int(m.group(1))
      
    c = ResultSet.CRIT_PATTERN.search(ln)
    if c != None and c.group(1) != None:
      self.crit_j = int(c.group(1))

  def set_gc(self, ln):
    """ Search for GC pause, record maximal value """
    m = ResultSet.PAUSE_Y_PATTERN.search(ln)
    if m != None and m.group(1) != None:
      py = float(m.group(1))
      if py > self.max_gc_y:
        self.max_gc_y = py

    c = ResultSet.PAUSE_F_PATTERN.search(ln)
    if c != None and c.group(1) != None:
      pf = float(c.group(1))
      if pf > self.max_gc_f:
        self.max_gc_f = pf
 
  def from_line(self, ln):
    """ Parse the line, grep entry point """
    self.set_max_crit(ln)
    self.set_gc(ln)
 
  def __str__ (self): 
    return str(self.run_name)

#-------- Core processing
def should_process(name):
  """ Limit scope of grep using hardcoded filepatterns, shell pattern match """
  global FILES
  matched = [x for x in FILES if fnmatch.fnmatch(name, x)]
  if len(matched) > 0:
    return True
  return False

def grep_file(filename):
  """ Grep over the single file """
  global _results
  line_count = 0
  (prefix, run_name) = ResultSet.get_run_name(filename)
  if _prefix_filter == None or fnmatch.fnmatch(prefix, _prefix_filter):
    rset = _results.get(run_name, None)
    if rset == None:
      rset = ResultSet(prefix, run_name, filename)
    with open(filename, "r") as fd:
      for ln in fd:
        line_count += 1
        rset.from_line(ln)
  return

def do_grep(dirname):
  """ Grep over the directory, processing files one by one, exceptions are reported and ignored """
  for root, dirs, files in os.walk(dirname, topdown=True): #pylint: disable=unused-variable
    for name in files:
      fname = os.path.join(root, name)  
      try:
        if should_process(name):
          grep_file(fname)
      except Exception as ex:
        report_exception(fname, ex)

# ----- Reporting
def do_write_txt(results, ofd=sys.stdout): 
  """ Output run results to text file or screen. If we are writing to file, options.sh file is included """
  global _show_zeros

  ofd.write("%-30s %-16s %-35s %-8s %-8s %-9s %-9s %-6s %-6s\n" % 
             ("Prefix", "Run", "Comments", "Max", "Crit", "GC_F", "GC_Y", "Pages", "GC_params"))

  row = 1 
  for r in results:
    if _show_zeros == True or r.max_j != 0:  
      opts = r.read_options()
      text1 = "%s/%s" % (opts.get("groups", "NA"), opts.get("pages", "NA"))
      text2 = "%s/%s/%s" % (opts.get("SurvivorRatio", "NA"), opts.get("TargetSurvivorRatio", "NA"), opts.get("MaxTenuringThreshold","NA"))
      comments = "\"" + r.read_description() + "\""
      ofd.write("%-30s %-16s %-35s %8d %6d %9.3f %9.3f %s %s\n" % 
                 (r.prefix, r.run_name, comments, r.max_j, r.crit_j, r.max_gc_f, r.max_gc_y, text1, text2))
      row += 1             
    if _top != None and row > _top:
      """ Write only N topmost items """
      break  

def do_write_xls(results, sc_name):
  """ Output run results to excel file """
  global _show_zeros, _cwd

  def set_row_bg(sheet, row, maxcol):
    for rows in sheet.iter_rows(min_row=row, max_row=row, min_col=1, max_col=maxcol):
      for cell in rows:
        cell.fill = PatternFill(fgColor="B6F2F2", fill_type = "solid")
    return

  def print_row(sheet, row, obj):
    opts = obj.read_options()
    comments = obj.read_description()

    row_values = (obj.prefix, "", comments, obj.max_j, obj.crit_j,
                  obj.max_gc_f, obj.max_gc_y, 
                  int(opts.get("groups", 0)), int(opts.get("pages", 0)),
                  int(opts.get("SurvivorRatio", 0)), int(opts.get("TargetSurvivorRatio", 0)),
                  int(opts.get("MaxTenuringThreshold", 0)))
    sheet.append(row_values)              
    # Setup clickable run_name
    sheet.cell(row=row, column= 2).value = obj.run_name
    sheet.cell(row=row, column= 2).hyperlink = "#'%s'!A1" % obj.run_name[:30]
    sheet.cell(row=row, column= 2).font = Font(color="0000CC", underline="single")

  def print_config(sh, obj):
    sh.cell(row=1, column=1).value = obj.options_file[len(_cwd):]

    (row, maxcol) = (2,4)

    sh.merge_cells(start_row=1, start_column=1, end_row=1, end_column=maxcol)
    set_row_bg(sh, 1, 4)
   
    sh.column_dimensions["A"].width = 25
    sh.column_dimensions["B"].width = 40

    with open(obj.options_file, 'r') as fd:
      for ln in fd:
        sh.append(ln.split(":", 1))
        row += 1

    if  obj.sysopts_file is not None:
      sh.cell(row=row, column=1).value = obj.sysopts_file[len(_cwd):]
      set_row_bg(sh, row, maxcol)

      n_row = row+1
      with open(obj.sysopts_file, 'r') as fd:
        for ln in fd:
          sh.append((ln,))
          n_row += 1
      for s_row in range(row, n_row):    
        sh.merge_cells(start_row=s_row, start_column=1, end_row=s_row, end_column=maxcol)
    return


  # Body of the function
  workbook = Workbook()
  ws = workbook.active
  ws.title = "Results"

  # Write table header
  ws.append(("Prefix", "Run", "Comments", "Max", "Crit", "GC F", "GC Y", "Groups", "Pages", "SR", "TSR", "MaxTT"))
  ws.column_dimensions["B"].width = 30
  ws.column_dimensions["C"].width = 50
  set_row_bg(ws, 1, 12)
  
  # Write table body and store config to a separate sheet 
  row = 2
  for r in results:
    if _show_zeros == True or r.max_j != 0:
      print_row(ws, row, r)
      if r.options_file is not None:
        print_config(workbook.create_sheet(r.run_name[:30], row), r)
      row += 1
    if _top != None and row > _top+1:
      """ Write only N topmost items """
      break  

  workbook.save(filename=sc_name)

def usage(msg=None):
  global HELP

  if msg is not None:
    print ("Error: %s" % msg)

  print(HELP)
  sys.exit(7)


if __name__ == '__main__':

  signal.signal(signal.SIGINT, signal_handler)

  try:
    opts, args = getopt.getopt(sys.argv[1:],
                                "ho:p:s:t:z",
                               ["help", "output", "prefix," "sort", "top", "zeros"])
  except getopt.GetoptError as ex:
    usage(ex)

  try:
    for o, a in opts:
      if o in ("-o", "--output"):
        _out_name = a
      elif o in ("-p", "--prefix"):
        _prefix_filter = a 
      elif o in ("-s", "--sort"):
        _sort = a 
      elif o in ("-t", "--top"):
        _top = int(a) 
      elif o in ("-z", "--zeros"):
        _show_zeros = not _show_zeros
      elif o in ("-h", "--help"):
        usage()
      else:
        assert False, "unhandled option"

    """ Validate parameters before actual run """
    assert _sort in ["prefix", "name", "max", "crit" ], "Invalid sort order, should be prefix, name, max, crit"
    if _out_name is not None:
      (filename, ext) = os.path.splitext(_out_name)
      assert ext in [".xlsx", ".txt"], "Unsupported output format '%s' should be xlsx, txt" % ext
      assert ext != ".xlsx" or _xls_possible, "Can't write to xls, install openpyxl"
  except Exception as ex:
    usage(ex)  

  dirlist = ["." ] if len(args) == 0 else args
  for dirname in dirlist:
    do_grep(dirname)  

  results = list(_results.values())
  if _sort == "name":
    results.sort(key=lambda x: x.run_name, reverse=True)
  elif _sort == "prefix":
    results.sort(key=lambda x: x.prefix, reverse=True)
  elif _sort == "max":
    results.sort(key=lambda x: x.max_j, reverse=True)
  elif _sort == "crit":
    results.sort(key=lambda x: x.crit_j, reverse=True)

  do_write_txt(results)

  if _out_name is not None:
    if ext == ".xlsx":
      do_write_xls(results, _out_name)
    elif ext == ".txt":
      with open(_out_name, "w") as ofd:
        do_write_txt(results, ofd)
    else:
      assert False, "Invalid output format %s" % _out_name  
        
